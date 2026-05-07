"""
AMIGO Blinds Calculator — Telegram Bot
"""
import asyncio, copy, json, logging, math, os, re, tempfile
from pathlib import Path

# Токен читается из переменной окружения BOT_TOKEN (Fly.io secret).
# Для локального запуска можно вписать сюда или задать переменную окружения.
TOKEN = ""

import aiohttp
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    ConversationHandler, filters, ContextTypes
)

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
log = logging.getLogger(__name__)

# ───────────────────────── files ──────────────────────────
CONFIG_FILE = Path("config.json")
PRICE_FILE  = Path("price_data.json")

# ─────────────────── conversation states ──────────────────
MECH_L, MECH_P, CAT, DIMS = range(4)

# ──────────────────── default config ──────────────────────
DEFAULT_CFG = {
    "rate_mode":   "auto",  # "auto" | "fixed"
    "cbr_rate":    None,
    "markup_pct":  5.0,
    "fixed_rate":  85.0,
    "final_rate":  None,
    "manual_disc": None,    # None = auto levels, number = fixed % override
    "last_mech_l": None,    # remembered last selection
    "last_mech_p": None,
    "last_cat":    None,
    "discount_levels": [
        {"min": 0,      "max": 15000,   "pct": 0},
        {"min": 15000,  "max": 30000,   "pct": 5},
        {"min": 30000,  "max": 50000,   "pct": 8},
        {"min": 50000,  "max": 75000,   "pct": 10},
        {"min": 75000,  "max": 125000,  "pct": 15},
        {"min": 125000, "max": 250000,  "pct": 18},
        {"min": 250000, "max": None,    "pct": 21},
    ],
}

# ─────────────────── config helpers ───────────────────────
def load_cfg() -> dict:
    if CONFIG_FILE.exists():
        saved = json.loads(CONFIG_FILE.read_text("utf-8"))
        cfg = copy.deepcopy(DEFAULT_CFG)   # deep copy — avoids mutating the default
        cfg.update(saved)
        return cfg
    return copy.deepcopy(DEFAULT_CFG)

def save_cfg(cfg: dict):
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), "utf-8")

def load_prices() -> dict:
    if PRICE_FILE.exists():
        return json.loads(PRICE_FILE.read_text("utf-8"))
    return {}

def save_prices(data: dict):
    PRICE_FILE.write_text(json.dumps(data, ensure_ascii=False), "utf-8")

# ─────────────────── price lookup ─────────────────────────
def nearest_idx(arr, val):
    # Ceiling: smallest table value >= actual (standard for blinds price sheets).
    # Each column/row means "blind up to X m", so always round UP to the next step.
    for i, v in enumerate(arr):
        if v >= val - 1e-9:
            return i
    return len(arr) - 1  # larger than table max → use biggest available

def uni_width(mech: str, w: float) -> float:
    """Only UNI 1 (not UNI 2): subtract 4 cm, snap to 10-cm step with 4-cm threshold.
    rem <= 4 → lower step; rem > 4 → upper step.
    48 cm → eff 44 → rem 4 ≤ 4 → 0.40 m
    48.1 cm → eff 44.1 → rem 4.1 > 4 → 0.50 m
    """
    if not re.search(r'uni\s*1', mech, re.I):
        return w
    eff = w * 100 - 4
    base = math.floor(eff / 10) * 10
    return (base if eff - base <= 4.0 + 1e-9 else base + 10) / 100

def lookup_price(mech, cat, w, h, prices):
    tbl = prices.get(mech, {}).get(cat)
    if not tbl:
        return None
    wi = nearest_idx(tbl["widths"], w)
    hi = nearest_idx(tbl["heights"], h)
    row = tbl["prices"][hi] if hi < len(tbl["prices"]) else None
    if row and wi < len(row) and row[wi] is not None:
        return {"price": float(row[wi]), "mw": tbl["widths"][wi], "mh": tbl["heights"][hi]}
    return None

# ─────────────────── discount ──────────────────────────────
def disc_pct(total_rub, levels):
    for lv in reversed(levels):
        if total_rub >= lv["min"]:
            return lv["pct"]
    return 0

def get_final_rate(cfg):
    if cfg["rate_mode"] == "fixed":
        return cfg["fixed_rate"]
    if cfg["cbr_rate"]:
        return round(cfg["cbr_rate"] * (1 + cfg["markup_pct"] / 100), 4)
    return cfg.get("final_rate") or cfg["fixed_rate"]

# ─────────────────── CBR rate ──────────────────────────────
async def fetch_cbr():
    url = "https://www.cbr-xml-daily.ru/daily_json.js"
    try:
        async with aiohttp.ClientSession() as s:
            async with s.get(url, timeout=aiohttp.ClientTimeout(total=6)) as r:
                data = await r.json(content_type=None)
                return data["Valute"]["USD"]["Value"]
    except Exception:
        return None

# ─────────────────── text parser ──────────────────────────
_DIM_RE = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*[xхXХ×]\s*(\d+(?:[.,]\d+)?)\s*([лпЛПlpLP])\.?\s*(.*)"
)

def parse_dims(text):
    items = []
    for line in text.splitlines():
        m = _DIM_RE.search(line.strip())
        if not m:
            continue
        w = float(m.group(1).replace(",", ".")) / 100
        h = float(m.group(2).replace(",", ".")) / 100
        t = "л" if m.group(3).lower() in ("л", "l") else "п"
        items.append({"w": w, "h": h, "type": t, "desc": m.group(4).strip()})
    return items

# ─────────────────── format helpers ───────────────────────
def fmt2(v):  return f"{v:.2f}"
def fmtrub(v): return f"{v:,.0f} ₽".replace(",", " ")

# ─────────────────── keyboards ────────────────────────────
MAIN_KB = ReplyKeyboardMarkup(
    [["📋 Рассчитать", "⚙️ Настройки"],
     ["📊 Обновить прайс", "💰 Курс USD"]],
    resize_keyboard=True,
)

def mechs_kb(prices):
    mechs = list(prices.keys())
    rows = []
    for i in range(0, len(mechs), 2):
        row = [InlineKeyboardButton(m, callback_data=f"mech|{m}") for m in mechs[i:i+2]]
        rows.append(row)
    return InlineKeyboardMarkup(rows)

def cats_kb(mech, prices):
    cats = sorted(prices.get(mech, {}).keys(),
                  key=lambda x: (x not in ("E", "Е"), int(x) if x.isdigit() else 0))
    rows = []
    for i in range(0, len(cats), 3):
        row = [InlineKeyboardButton(f"Кат. {c}", callback_data=f"cat|{c}") for c in cats[i:i+3]]
        rows.append(row)
    return InlineKeyboardMarkup(rows)

# ═══════════════════ /start ═══════════════════════════════
async def cmd_start(update: Update, _):
    await update.message.reply_text(
        "👋 *AMIGO — Калькулятор жалюзей*\n\nВыберите действие:",
        parse_mode="Markdown", reply_markup=MAIN_KB,
    )

# ═══════════════════ MAIN MENU ════════════════════════════
async def menu(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    t = update.message.text
    if   t == "📋 Рассчитать":   return await calc_start(update, ctx)
    elif t == "⚙️ Настройки":   return await settings(update, ctx)
    elif t == "📊 Обновить прайс": await update.message.reply_text("📎 Отправьте файл .xlsx с прайсом:")
    elif t == "💰 Курс USD":     return await rate_info(update, ctx)

# ═══════════════════ CALC FLOW ════════════════════════════
async def calc_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    prices = load_prices()
    if not prices:
        await update.message.reply_text("❌ Прайс не загружен.\nОтправьте файл через «📊 Обновить прайс».")
        return ConversationHandler.END
    ctx.user_data["prices"] = prices
    cfg = load_cfg()
    last_l = cfg.get("last_mech_l")
    hint = f"\n_Последний: {last_l}_" if last_l else ""
    await update.message.reply_text(
        f"1️⃣ Механизм для *«Л»* (лента):{hint}", parse_mode="Markdown",
        reply_markup=mechs_kb(prices),
    )
    return MECH_L

async def got_mech_l(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    mech = q.data.split("|", 1)[1]
    ctx.user_data["mech_l"] = mech
    cfg = load_cfg(); cfg["last_mech_l"] = mech; save_cfg(cfg)
    last_p = cfg.get("last_mech_p")
    hint = f"\n_Последний: {last_p}_" if last_p else ""
    await q.edit_message_text(
        f"✅ Механизм Л: *{mech}*\n\n2️⃣ Механизм для *«П»* (полотно):{hint}",
        parse_mode="Markdown", reply_markup=mechs_kb(ctx.user_data["prices"]),
    )
    return MECH_P

async def got_mech_p(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    mech = q.data.split("|", 1)[1]
    ctx.user_data["mech_p"] = mech
    cfg = load_cfg(); cfg["last_mech_p"] = mech; save_cfg(cfg)
    ref = ctx.user_data["mech_l"]
    last_c = cfg.get("last_cat")
    hint = f"\n_Последняя: {last_c}_" if last_c else ""
    await q.edit_message_text(
        f"✅ Механизм П: *{mech}*\n\n3️⃣ Категория:{hint}",
        parse_mode="Markdown",
        reply_markup=cats_kb(ref, ctx.user_data["prices"]),
    )
    return CAT

async def got_cat(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    cat = q.data.split("|", 1)[1]
    ctx.user_data["cat"] = cat
    cfg = load_cfg(); cfg["last_cat"] = cat; save_cfg(cfg)
    await q.edit_message_text(
        f"✅ Категория: *{cat}*\n\n"
        "4️⃣ Введите размеры (каждая строка — позиция):\n\n"
        "```\n49,7 x 138 л  Уни 1 белый\n50,2 x 128 п  Аура V-14\n62,3 x 138 л\n```",
        parse_mode="Markdown",
    )
    return DIMS

async def got_dims(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    items = parse_dims(update.message.text)
    if not items:
        await update.message.reply_text(
            "❌ Не распознано. Формат: `49,7 x 138 л` или `50 x 128 п`",
            parse_mode="Markdown",
        )
        return DIMS

    cfg    = load_cfg()
    prices = ctx.user_data["prices"]
    mL, mP = ctx.user_data["mech_l"], ctx.user_data["mech_p"]
    cat    = ctx.user_data["cat"]
    rate   = get_final_rate(cfg)

    if not rate:
        await update.message.reply_text("❌ Курс не установлен. Зайдите в «💰 Курс USD».")
        return ConversationHandler.END

    lines, total_usd, errors = [], 0.0, 0

    for i, item in enumerate(items, 1):
        mech = mL if item["type"] == "л" else mP
        adj_w = uni_width(mech, item["w"])
        res   = lookup_price(mech, cat, adj_w, item["h"], prices)
        if not res:
            w_cm = item["w"] * 100
            h_cm = item["h"] * 100
            lines.append(f"⚠️ {i}. Цена не найдена — {w_cm:.1f}×{h_cm:.1f}см {item['type'].upper()}")
            errors += 1
            continue
        usd = round(res["price"], 2)
        rub = round(usd * rate, 2)
        total_usd += usd
        uni_tag = " *(−4см)*" if re.search(r'uni\s*1', mech, re.I) else ""
        desc = f" {item['desc']}" if item["desc"] else ""
        lines.append(
            f"{i}. *{mech}*{uni_tag} {res['mw']}×{res['mh']}м {item['type'].upper()}{desc}\n"
            f"   {fmt2(usd)}$ × {fmt2(rate)} = *{fmtrub(rub)}*"
        )

    total_usd     = round(total_usd, 2)
    total_rub_raw = round(total_usd * rate, 2)   # single multiply — no accumulation error
    manual   = cfg.get("manual_disc")
    disc     = manual if manual is not None else disc_pct(total_rub_raw, cfg["discount_levels"])
    disc_amt = round(total_rub_raw * disc / 100, 2)
    total    = round(total_rub_raw - disc_amt, 2)
    ok       = len(items) - errors

    sep    = "─" * 28
    footer = (
        f"\n{sep}\n"
        f"💵 Сумма: *{fmt2(total_usd)} $*\n"
        f"💴 В рублях: {fmtrub(total_rub_raw)}\n"
        + (f"🏷 Скидка {disc}%: −{fmtrub(disc_amt)}\n" if disc else "")
        + f"✅ *Итого: {fmtrub(total)}*"
    )
    header = f"📋 *Расчёт* — {ok} поз. | курс {fmt2(rate)} ₽/$\n{sep}\n"

    body = "\n".join(lines)
    msg  = header + body + footer

    # If too long — truncate items but always keep footer
    if len(msg) > 4096:
        max_body = 4096 - len(header) - len(footer) - 60
        cut = body.rfind("\n", 0, max_body)
        if cut < 0:
            cut = max_body
        body = body[:cut] + f"\n\n_...ещё позиции скрыты (показаны {body[:cut].count(chr(10))+1} из {ok})_"
        msg  = header + body + footer
        if len(msg) > 4096:
            msg = msg[:4090] + "…"

    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=MAIN_KB)
    return ConversationHandler.END

async def cancel(update: Update, _):
    await update.message.reply_text("Отменено.", reply_markup=MAIN_KB)
    return ConversationHandler.END

# ═══════════════════ RATE ════════════════════════════════
async def rate_info(update: Update, _: ContextTypes.DEFAULT_TYPE):
    cfg  = load_cfg()
    rate = get_final_rate(cfg)
    mode = "🔒 Фиксированный" if cfg["rate_mode"] == "fixed" else "⚡ Авто (ЦБ)"
    txt  = (
        f"💰 *Курс USD*\n\n"
        f"Режим: {mode}\n"
        f"ЦБ РФ: {fmt2(cfg['cbr_rate']) if cfg['cbr_rate'] else '—'} ₽\n"
        f"Наценка: {cfg['markup_pct']}%\n"
        f"*Итоговый курс: {fmt2(rate)} ₽/$*"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔄 Обновить с ЦБ РФ",    callback_data="rate|refresh")],
        [InlineKeyboardButton("✏️ Ввести курс вручную", callback_data="rate|manual")],
        [InlineKeyboardButton("📊 Изменить наценку %",  callback_data="rate|markup")],
    ])
    msg = update.message or update.callback_query.message
    await msg.reply_text(txt, parse_mode="Markdown", reply_markup=kb)

async def rate_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    action = q.data.split("|", 1)[1]
    cfg = load_cfg()

    if action == "refresh":
        cbr = await fetch_cbr()
        if cbr:
            cfg["cbr_rate"]  = cbr
            cfg["rate_mode"] = "auto"
            cfg["final_rate"] = round(cbr * (1 + cfg["markup_pct"] / 100), 4)
            save_cfg(cfg)
            await q.edit_message_text(
                f"✅ Курс ЦБ: *{fmt2(cbr)} ₽/$*\n"
                f"С наценкой {cfg['markup_pct']}%: *{fmt2(cfg['final_rate'])} ₽/$*",
                parse_mode="Markdown",
            )
        else:
            await q.edit_message_text("❌ Не удалось получить курс ЦБ РФ. Попробуйте позже.")

    elif action == "manual":
        ctx.user_data["awaiting"] = "fixed_rate"
        await q.edit_message_text(
            "✏️ Введите курс числом, например `87.50`:", parse_mode="Markdown"
        )

    elif action == "markup":
        ctx.user_data["awaiting"] = "markup_pct"
        await q.edit_message_text(
            f"📊 Текущая наценка: *{cfg['markup_pct']}%*\n\nВведите новый процент:",
            parse_mode="Markdown",
        )

# ═══════════════════ SETTINGS (скидки) ═══════════════════
async def settings(update: Update, _):
    cfg    = load_cfg()
    levels = cfg["discount_levels"]
    manual = cfg.get("manual_disc")

    lines  = []
    for i, lv in enumerate(levels):
        max_s = f"{int(lv['max']):,}".replace(",", " ") if lv["max"] else "∞"
        min_s = f"{int(lv['min']):,}".replace(",", " ")
        lines.append(f"{i+1}. {min_s} – {max_s} ₽  →  *{lv['pct']}%*")

    if manual is not None:
        disc_status = f"🔒 *Ручная скидка: {manual}%* (авто-уровни отключены)"
    else:
        disc_status = "⚡ Скидка: авто по уровням"

    kb_rows = []
    for i in range(0, len(levels), 3):
        kb_rows.append([
            InlineKeyboardButton(f"✏️ Уровень {j+1}", callback_data=f"disc|{j}")
            for j in range(i, min(i+3, len(levels)))
        ])
    manual_row = [InlineKeyboardButton("✏️ Скидка вручную %", callback_data="discm|set")]
    if manual is not None:
        manual_row.append(InlineKeyboardButton("🔄 Вернуть авто", callback_data="discm|clear"))
    kb_rows.append(manual_row)

    msg = update.message or update.callback_query.message
    await msg.reply_text(
        f"⚙️ *Настройки скидок*\n\n{disc_status}\n\n"
        "─── Авто-уровни ───\n" + "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb_rows),
    )

async def disc_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    idx = int(q.data.split("|", 1)[1])
    cfg = load_cfg()
    lv  = cfg["discount_levels"][idx]
    max_s = f"{int(lv['max']):,}".replace(",", " ") if lv["max"] else "∞"
    ctx.user_data["disc_idx"] = idx
    await q.edit_message_text(
        f"✏️ *Уровень {idx+1}*\n"
        f"От: {int(lv['min']):,} ₽\n".replace(",", " ") +
        f"До: {max_s} ₽\n"
        f"Скидка: *{lv['pct']}%*\n\n"
        "Что изменить?",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("% скидки",    callback_data=f"discf|pct|{idx}"),
             InlineKeyboardButton("Порог от, ₽", callback_data=f"discf|min|{idx}")],
        ]),
    )

async def disc_field_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    _, field, idx = q.data.split("|")
    ctx.user_data["disc_idx"]   = int(idx)
    ctx.user_data["disc_field"] = field
    hint = "процент скидки, например `10`" if field == "pct" else "порог от (в ₽), например `50000`"
    ctx.user_data["awaiting"] = "disc_val"
    await q.edit_message_text(f"Введите {hint}:", parse_mode="Markdown")

async def disc_manual_cb(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query; await q.answer()
    action = q.data.split("|", 1)[1]
    cfg = load_cfg()
    if action == "clear":
        cfg["manual_disc"] = None
        save_cfg(cfg)
        await q.edit_message_text("✅ Скидка переключена на *авто* (по уровням)", parse_mode="Markdown")
    elif action == "set":
        ctx.user_data["awaiting"] = "manual_disc"
        await q.edit_message_text(
            "✏️ Введите скидку в % (0–100), например `10`:\n\n"
            "_Это значение будет применяться ко всем расчётам вместо авто-уровней._",
            parse_mode="Markdown",
        )

# ═══════════════════ TEXT INPUT ══════════════════════════
async def text_input(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    awaiting = ctx.user_data.get("awaiting")
    text = update.message.text.strip().replace(",", ".")

    if awaiting == "fixed_rate":
        try:
            val = float(text)
            cfg = load_cfg()
            cfg.update({"fixed_rate": val, "rate_mode": "fixed", "final_rate": val})
            save_cfg(cfg)
            ctx.user_data.pop("awaiting", None)
            await update.message.reply_text(
                f"✅ Курс зафиксирован: *{fmt2(val)} ₽/$*",
                parse_mode="Markdown", reply_markup=MAIN_KB,
            )
        except ValueError:
            await update.message.reply_text("❌ Введите число, например `87.50`", parse_mode="Markdown")

    elif awaiting == "markup_pct":
        try:
            val = float(text)
            cfg = load_cfg()
            cfg["markup_pct"] = val
            if cfg["cbr_rate"]:
                cfg["final_rate"] = round(cfg["cbr_rate"] * (1 + val / 100), 4)
            save_cfg(cfg)
            ctx.user_data.pop("awaiting", None)
            await update.message.reply_text(
                f"✅ Наценка: *{val}%*", parse_mode="Markdown", reply_markup=MAIN_KB,
            )
        except ValueError:
            await update.message.reply_text("❌ Введите число, например `5`", parse_mode="Markdown")

    elif awaiting == "disc_val":
        try:
            val   = float(text)
            idx   = ctx.user_data["disc_idx"]
            field = ctx.user_data["disc_field"]
            cfg   = load_cfg()
            cfg["discount_levels"][idx][field] = val
            save_cfg(cfg)
            ctx.user_data.pop("awaiting", None)
            label = "% скидки" if field == "pct" else "порог от"
            await update.message.reply_text(
                f"✅ Уровень {idx+1}, {label} обновлён: *{val}*",
                parse_mode="Markdown", reply_markup=MAIN_KB,
            )
        except ValueError:
            await update.message.reply_text("❌ Введите число", parse_mode="Markdown")

    elif awaiting == "manual_disc":
        try:
            val = float(text)
            if not (0 <= val <= 100):
                raise ValueError
            cfg = load_cfg()
            cfg["manual_disc"] = val
            save_cfg(cfg)
            ctx.user_data.pop("awaiting", None)
            await update.message.reply_text(
                f"✅ Ручная скидка установлена: *{val}%*\n"
                "_(Авто-уровни отключены. Чтобы вернуть авто — ⚙️ Настройки → 🔄 Вернуть авто)_",
                parse_mode="Markdown", reply_markup=MAIN_KB,
            )
        except ValueError:
            await update.message.reply_text("❌ Введите число от 0 до 100, например `10`", parse_mode="Markdown")

    else:
        await menu(update, ctx)

# ═══════════════════ EXCEL UPLOAD ════════════════════════
def _pn(v):
    if v is None: return float("nan")
    if isinstance(v, (int, float)): return float(v)
    return float(str(v).replace(",", ".").replace(" ", "").replace("\xa0", "") or "nan")

def parse_excel(path: str) -> dict:
    import openpyxl
    wb  = openpyxl.load_workbook(path, data_only=True)
    res = {}

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if len(rows) < 3:
            continue

        mech_data = {}

        for ri, row in enumerate(rows):
            for ci, cell in enumerate(row):
                if not cell or "категория" not in str(cell).lower():
                    continue
                # Determine category letter
                s  = str(cell).replace("Е", "E").replace("е", "e")
                m  = re.search(r"([E1-9])\s*[кk]", s, re.I) or re.search(r"[кk]атегория\s*([E1-9])", s, re.I)
                if not m:
                    continue
                cat = m.group(1).upper()

                # Find width row (numbers 0.3–5.5) in next 3 rows
                w_row, w_cols, widths = -1, [], []
                for dr in range(1, 4):
                    if ri + dr >= len(rows): break
                    nr = rows[ri + dr]
                    tw, tc, started = [], [], False
                    for cc in range(ci, len(nr)):
                        n = _pn(nr[cc])
                        if not (n != n) and 0.3 <= n <= 5.5:
                            tw.append(round(n * 10) / 10); tc.append(cc); started = True
                        elif started:
                            break
                    if len(tw) >= 3:
                        w_row = ri + dr; widths = tw; w_cols = tc; break

                if w_row == -1 or not w_cols:
                    continue

                h_col = w_cols[0] - 1
                if h_col < 0:
                    continue

                heights, prices = [], []
                for pr in range(w_row + 1, len(rows)):
                    r2 = rows[pr]
                    hv = _pn(r2[h_col]) if h_col < len(r2) else float("nan")
                    if hv != hv or not (0.3 <= hv <= 6.0):
                        break
                    heights.append(round(hv * 10) / 10)
                    price_row = []
                    for cc in w_cols:
                        p = _pn(r2[cc]) if cc < len(r2) else float("nan")
                        price_row.append(None if p != p else p)
                    prices.append(price_row)

                if heights and prices:
                    mech_data[cat] = {"widths": widths, "heights": heights, "prices": prices}

        if mech_data:
            res[sheet_name] = mech_data

    return res

async def handle_doc(update: Update, _):
    doc = update.message.document
    if not doc.file_name.lower().endswith((".xlsx", ".xls")):
        await update.message.reply_text("❌ Нужен файл .xlsx")
        return
    await update.message.reply_text("⏳ Обрабатываю…")
    path = Path(tempfile.gettempdir()) / f"{doc.file_unique_id}.xlsx"
    try:
        f = await doc.get_file()
        await f.download_to_drive(str(path))
        data = parse_excel(path)
        if not data:
            await update.message.reply_text("❌ Не удалось распознать формат прайса.")
            return
        save_prices(data)
        mechs = ", ".join(data.keys())
        await update.message.reply_text(
            f"✅ Прайс обновлён: *{len(data)}* мех.\n_{mechs}_",
            parse_mode="Markdown", reply_markup=MAIN_KB,
        )
    except Exception as e:
        log.exception("Excel parse error")
        await update.message.reply_text(f"❌ Ошибка: {e}")
    finally:
        path.unlink(missing_ok=True)   # always clean up temp file

async def error_handler(_update: object, ctx: ContextTypes.DEFAULT_TYPE):
    log.error("Unhandled exception", exc_info=ctx.error)

# ═══════════════════ MAIN ════════════════════════════════
def main():
    # Prevent httpcore/httpx from auto-detecting the Windows system proxy,
    # which causes ConnectTimeout / RemoteProtocolError on every Telegram API call.
    for _var in ("HTTP_PROXY", "HTTPS_PROXY", "http_proxy", "https_proxy",
                 "ALL_PROXY", "all_proxy", "NO_PROXY", "no_proxy"):
        os.environ.pop(_var, None)

    token = TOKEN or os.environ.get("BOT_TOKEN", "")
    if not token:
        raise RuntimeError("Вставьте токен в переменную TOKEN в начале bot.py")

    async def on_startup(_):
        cfg = load_cfg()
        if cfg["rate_mode"] == "auto":
            cbr = await fetch_cbr()
            if cbr:
                cfg["cbr_rate"]   = cbr
                cfg["final_rate"] = round(cbr * (1 + cfg["markup_pct"] / 100), 4)
                save_cfg(cfg)
                log.info("CBR rate loaded on startup: %.4f", cfg["final_rate"])

    app = (
        Application.builder()
        .token(token)
        .proxy(None)
        .get_updates_proxy(None)
        .post_init(on_startup)
        .build()
    )

    calc_conv = ConversationHandler(
        entry_points=[MessageHandler(filters.Regex("^📋 Рассчитать$"), calc_start)],
        states={
            MECH_L: [CallbackQueryHandler(got_mech_l, pattern="^mech\\|")],
            MECH_P: [CallbackQueryHandler(got_mech_p, pattern="^mech\\|")],
            CAT:    [CallbackQueryHandler(got_cat,    pattern="^cat\\|")],
            DIMS:   [MessageHandler(filters.TEXT & ~filters.COMMAND, got_dims)],
        },
        fallbacks=[CommandHandler("cancel", cancel),
                   CommandHandler("start",  cancel)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(calc_conv)
    app.add_handler(CallbackQueryHandler(rate_cb,        pattern="^rate\\|"))
    app.add_handler(CallbackQueryHandler(disc_cb,        pattern="^disc\\|"))
    app.add_handler(CallbackQueryHandler(disc_field_cb,  pattern="^discf\\|"))
    app.add_handler(CallbackQueryHandler(disc_manual_cb, pattern="^discm\\|"))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_doc))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_input))
    app.add_error_handler(error_handler)

    log.info("Bot started")
    app.run_polling()

if __name__ == "__main__":
    import asyncio, time
    delay = 5
    while True:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            main()
            break  # чистый выход (не крэш)
        except KeyboardInterrupt:
            log.info("Bot stopped by user")
            break
        except Exception as e:
            log.error("Bot crashed: %s — restart in %ds", e, delay)
            time.sleep(delay)
            delay = min(delay * 2, 60)  # 5 → 10 → 20 → 40 → 60s
